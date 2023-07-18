import importlib
import subprocess

# List of required dependencies
REQUIRED_PACKAGES = [
    'flask',
    'openpyxl'
]

# Check and install missing dependencies
for package in REQUIRED_PACKAGES:
    try:
        importlib.import_module(package)
    except ImportError:
        subprocess.check_call(['pip', 'install', package])

# Import the required modules
from flask import Flask, render_template, request, send_from_directory, redirect, send_file
import os
import re
from collections import defaultdict
from openpyxl import Workbook

app = Flask(__name__)

def create_excel_file(file_path, sheet_data, most_sold_items=None, most_bought_items=None, item_limit=None, player_limit=None):
    # Delete existing file if it exists
    if os.path.exists(file_path):
        os.remove(file_path)

    wb = Workbook()

    for sheet_name, day_data in sheet_data.items():
        sheet = wb.create_sheet(title=sheet_name)
        sheet["A1"] = "Date"
        sheet["B1"] = "Player"
        sheet["C1"] = "Amount"

        row = 2
        for date, players in day_data.items():
            sheet.cell(row=row, column=1).value = date
            col = 2
            for player, amount in players:
                sheet.cell(row=row, column=col).value = player
                sheet.cell(row=row, column=col + 1).value = amount
                col += 2
            row += 1

        # Adjust cell width for the sheet
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.coordinate == f"{column}1":
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

    if most_sold_items:
        sold_sheet = wb.create_sheet(title="Most Sold Items")
        sold_sheet["A1"] = "Date"
        sold_sheet["B1"] = "Most Sold Items"

        row = 2
        for date, items in most_sold_items.items():
            sold_sheet[f"A{row}"] = date
            if item_limit:
                items = items[:item_limit]
            col = 2
            for item in items:
                sold_sheet.cell(row=row, column=col).value = item
                col += 1
            row += 1

        # Adjust cell width for the most sold items sheet
        for column_cells in sold_sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.coordinate == f"{column}1":
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sold_sheet.column_dimensions[column].width = adjusted_width

    if most_bought_items:
        bought_sheet = wb.create_sheet(title="Most Bought Items")
        bought_sheet["A1"] = "Date"
        bought_sheet["B1"] = "Most Bought Items"

        row = 2
        for date, items in most_bought_items.items():
            bought_sheet[f"A{row}"] = date
            if item_limit:
                items = items[:item_limit]
            col = 2
            for item in items:
                bought_sheet.cell(row=row, column=col).value = item
                col += 1
            row += 1

        # Adjust cell width for the most bought items sheet
        for column_cells in bought_sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.coordinate == f"{column}1":
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            bought_sheet.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    print(f"Excel file saved successfully at: {file_path}")


from collections import defaultdict

def find_most_sell_players_per_day(file_path, player_limit=None):
    sell_players_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "sold" in line.lower():
                date_info = re.search(r"\[(\d{4}-\d{2}-\d{2})\s\d{2}:\d{2}:\d{2}\]", line)
                player_info = re.search(r"\[\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\] - (\w+) sold", line)
                amount_info = re.search(r"for\s+\$([\d,\.]+)", line)
                if date_info and player_info and amount_info:
                    date = date_info.group(1)
                    player = player_info.group(1).strip()
                    amount = float(amount_info.group(1).replace(",", ""))
                    amount = round(amount, 2)  # Round up to 2 decimal places
                    sell_players_per_day[date].append((player, amount))

    sell_players_per_day_sorted = {}
    for date, players in sell_players_per_day.items():
        player_sum = defaultdict(float)
        for player, amount in players:
            player_sum[player] += amount

        players_sorted = sorted(player_sum.items(), key=lambda x: x[1], reverse=True)
        if player_limit:
            players_sorted = players_sorted[:player_limit]
        sell_players_per_day_sorted[date] = players_sorted

    return sell_players_per_day_sorted


def find_most_buy_players_per_day(file_path, player_limit=None):
    buy_players_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "bought" in line.lower():
                date_info = re.search(r"\[(\d{4}-\d{2}-\d{2})\s\d{2}:\d{2}:\d{2}\]", line)
                player_info = re.search(r"\[\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\] - (\w+) bought", line)
                amount_info = re.search(r"for\s+\$([\d,\.]+)", line)
                if date_info and player_info and amount_info:
                    date = date_info.group(1)
                    player = player_info.group(1).strip()
                    amount = float(amount_info.group(1).replace(",", ""))
                    amount = round(amount, 2)  # Round up to 2 decimal places
                    buy_players_per_day[date].append((player, amount))

    buy_players_per_day_sorted = {}
    for date, players in buy_players_per_day.items():
        player_sum = defaultdict(float)
        for player, amount in players:
            player_sum[player] += amount

        players_sorted = sorted(player_sum.items(), key=lambda x: x[1], reverse=True)
        if player_limit:
            players_sorted = players_sorted[:player_limit]
        buy_players_per_day_sorted[date] = players_sorted

    return buy_players_per_day_sorted



def find_most_sold_items_per_day(file_path, include_id=True):
    sold_items_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "sold" in line.lower():
                date = line.split()[0][1:]
                item_info = line.split("sold")[1].strip()
                item = extract_item_name(item_info)
                sold_items_per_day[date].append(item)

    most_sold_items_per_day = {}
    for date, items in sold_items_per_day.items():
        item_count = defaultdict(int)
        for item in items:
            item_count[item] += 1
        sorted_items = sorted(item_count.items(), key=lambda x: x[1], reverse=True)
        most_sold_items = []
        for item, count in sorted_items:
            if include_id:
                most_sold_items.append(f"{item}x{count}")
            else:
                item_name = item.split("(")[0].strip()
                most_sold_items.append(f"{item_name}x{count}")
        most_sold_items_per_day[date] = most_sold_items

    return most_sold_items_per_day


def find_most_bought_items_per_day(file_path, include_id=True):
    bought_items_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "bought" in line.lower():
                date = line.split()[0][1:]
                item_info = line.split("bought")[1].strip()
                item = extract_item_name(item_info)
                bought_items_per_day[date].append(item)

    most_bought_items_per_day = {}
    for date, items in bought_items_per_day.items():
        item_count = defaultdict(int)
        for item in items:
            item_count[item] += 1
        sorted_items = sorted(item_count.items(), key=lambda x: x[1], reverse=True)
        most_bought_items = []
        for item, count in sorted_items:
            if include_id:
                most_bought_items.append(f"{item}x{count}")
            else:
                item_name = item.split("(")[0].strip()
                most_bought_items.append(f"{item_name}x{count}")
        most_bought_items_per_day[date] = most_bought_items

    return most_bought_items_per_day


def extract_item_name(item_info):
    item_parts = item_info.split("x")
    if len(item_parts) < 2:
        return item_info.strip()

    item_name = "x".join(item_parts[1:]).strip().split("for")[0].strip()
    return item_name

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    log_file = request.files['logfile']
    item_limit = int(request.form['itemlimit'])
    player_limit = int(request.form['playerlimit'])
    include_item_id = 'includeid' in request.form

    # Save the uploaded log file to a temporary location
    log_file_path = os.path.join(app.config['UPLOAD_FOLDER'], log_file.filename)
    log_file.save(log_file_path)

    # Process the log file and generate the data
    most_sold_items = find_most_sold_items_per_day(log_file_path, include_id=include_item_id)
    most_bought_items = find_most_bought_items_per_day(log_file_path, include_id=include_item_id)
    most_sold_items_limited = {date: items[:item_limit] if item_limit else items for date, items in
                               most_sold_items.items()}
    most_bought_items_limited = {date: items[:item_limit] if item_limit else items for date, items in
                                most_bought_items.items()}

    buy_players_per_day = find_most_buy_players_per_day(log_file_path, player_limit=5)
    sell_players_per_day = find_most_sell_players_per_day(log_file_path, player_limit=5)

    sheet_data = {
        "Most Buy Players per Day": buy_players_per_day,
        "Most Sell Players per Day": sell_players_per_day
    }

    # Generate the Excel file
    excel_file_path = os.path.join(app.config['OUTPUT_FOLDER'], "transaction_summary.xlsx")
    create_excel_file(excel_file_path, sheet_data, most_sold_items_limited, most_bought_items_limited,
                      item_limit=item_limit, player_limit=player_limit)

    return render_template('results.html',
                           most_sold_items=most_sold_items_limited,
                           most_bought_items=most_bought_items_limited,
                           buy_players_per_day=buy_players_per_day,
                           sell_players_per_day=sell_players_per_day)

@app.route('/download')
def download():
    excel_file_path = os.path.join(app.config['OUTPUT_FOLDER'], "transaction_summary.xlsx")
    return send_file(excel_file_path, as_attachment=True, download_name="transaction_summary.xlsx")

if __name__ == '__main__':
    app.config['UPLOAD_FOLDER'] = 'uploads'  # Folder to store uploaded files
    app.config['OUTPUT_FOLDER'] = 'output'  # Folder to store generated Excel file
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
    app.run(port=8000)
