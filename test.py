import os
import re
from collections import defaultdict
from openpyxl import Workbook


def create_excel_file(file_path, sheet_data, most_sold_items=None, most_bought_items=None, item_limit=None):
    # Delete existing file if it exists
    if os.path.exists(file_path):
        os.remove(file_path)

    wb = Workbook()

    for sheet_name, day_data in sheet_data.items():
        sheet = wb.create_sheet(title=sheet_name)
        sheet["A1"] = "Date"
        sheet["B1"] = "Player"
        sheet["C1"] = "Amount"

        row = 2
        for date, players in day_data.items():
            sheet.cell(row=row, column=1).value = date
            col = 2
            for player, amount in players:
                sheet.cell(row=row, column=col).value = player
                sheet.cell(row=row, column=col + 1).value = amount
                row += 1

        # Adjust cell width for the sheet
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.coordinate == f"{column}1":
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

    if most_sold_items:
        sold_sheet = wb.create_sheet(title="Most Sold Items")
        sold_sheet["A1"] = "Date"
        sold_sheet["B1"] = "Most Sold Items"

        row = 2
        for date, items in most_sold_items.items():
            sold_sheet[f"A{row}"] = date
            if item_limit:
                items = items[:item_limit]
            col = 2
            for item in items:
                sold_sheet.cell(row=row, column=col).value = item
                col += 1
            row += 1

        # Adjust cell width for the most sold items sheet
        for column_cells in sold_sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.coordinate == f"{column}1":
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sold_sheet.column_dimensions[column].width = adjusted_width

    if most_bought_items:
        bought_sheet = wb.create_sheet(title="Most Bought Items")
        bought_sheet["A1"] = "Date"
        bought_sheet["B1"] = "Most Bought Items"

        row = 2
        for date, items in most_bought_items.items():
            bought_sheet[f"A{row}"] = date
            if item_limit:
                items = items[:item_limit]
            col = 2
            for item in items:
                bought_sheet.cell(row=row, column=col).value = item
                col += 1
            row += 1

        # Adjust cell width for the most bought items sheet
        for column_cells in bought_sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.coordinate == f"{column}1":
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            bought_sheet.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    print(f"Excel file saved successfully at: {file_path}")


def find_most_buy_players_per_day(file_path, player_limit=None):
    buy_players_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "bought" in line.lower():
                date_info = re.search(r"\[(\d{4}-\d{2}-\d{2})\s\d{2}:\d{2}:\d{2}\]", line)
                player_info = re.search(r"\[\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\] - (\w+) bought", line)
                amount_info = re.search(r"for\s+\$([\d,\.]+)", line)
                if date_info and player_info and amount_info:
                    date = date_info.group(1)
                    player = player_info.group(1).strip()
                    amount = float(amount_info.group(1).replace(",", ""))
                    buy_players_per_day[date].append((player, amount))

    buy_players_per_day_sorted = {}
    for date, players in buy_players_per_day.items():
        players_sorted = sorted(players, key=lambda x: x[1], reverse=True)
        if player_limit:
            players_sorted = players_sorted[:player_limit]
        buy_players_per_day_sorted[date] = players_sorted

    return buy_players_per_day_sorted


def find_most_sell_players_per_day(file_path, player_limit=None):
    sell_players_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "sold" in line.lower():
                date_info = re.search(r"\[(\d{4}-\d{2}-\d{2})\s\d{2}:\d{2}:\d{2}\]", line)
                player_info = re.search(r"\[\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\] - (\w+) sold", line)
                amount_info = re.search(r"for\s+\$([\d,\.]+)", line)
                if date_info and player_info and amount_info:
                    date = date_info.group(1)
                    player = player_info.group(1).strip()
                    amount = float(amount_info.group(1).replace(",", ""))
                    sell_players_per_day[date].append((player, amount))

    sell_players_per_day_sorted = {}
    for date, players in sell_players_per_day.items():
        players_sorted = sorted(players, key=lambda x: x[1], reverse=True)
        if player_limit:
            players_sorted = players_sorted[:player_limit]
        sell_players_per_day_sorted[date] = players_sorted

    return sell_players_per_day_sorted


def find_most_sold_items_per_day(file_path, include_id=True):
    sold_items_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "sold" in line.lower():
                date = line.split()[0][1:]
                item_info = line.split("sold")[1].strip()
                item = extract_item_name(item_info)
                sold_items_per_day[date].append(item)

    most_sold_items_per_day = {}
    for date, items in sold_items_per_day.items():
        item_count = defaultdict(int)
        for item in items:
            item_count[item] += 1
        sorted_items = sorted(item_count.items(), key=lambda x: x[1], reverse=True)
        most_sold_items = []
        for item, count in sorted_items:
            if include_id:
                most_sold_items.append(f"{item}x{count}")
            else:
                item_name = item.split("(")[0].strip()
                most_sold_items.append(f"{item_name}x{count}")
        most_sold_items_per_day[date] = most_sold_items

    return most_sold_items_per_day


def find_most_bought_items_per_day(file_path, include_id=True):
    bought_items_per_day = defaultdict(list)

    with open(file_path, 'r') as file:
        for line in file:
            if "bought" in line.lower():
                date = line.split()[0][1:]
                item_info = line.split("bought")[1].strip()
                item = extract_item_name(item_info)
                bought_items_per_day[date].append(item)

    most_bought_items_per_day = {}
    for date, items in bought_items_per_day.items():
        item_count = defaultdict(int)
        for item in items:
            item_count[item] += 1
        sorted_items = sorted(item_count.items(), key=lambda x: x[1], reverse=True)
        most_bought_items = []
        for item, count in sorted_items:
            if include_id:
                most_bought_items.append(f"{item}x{count}")
            else:
                item_name = item.split("(")[0].strip()
                most_bought_items.append(f"{item_name}x{count}")
        most_bought_items_per_day[date] = most_bought_items

    return most_bought_items_per_day


def extract_item_name(item_info):
    item_parts = item_info.split("x")
    if len(item_parts) < 2:
        return item_info.strip()

    item_name = "x".join(item_parts[1:]).strip().split("for")[0].strip()
    return item_name


def main():
    # Get the path of the script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Log file path (assuming it's in the same directory as the script)
    log_file = os.path.join(script_dir, "transaction-log.txt")

    # Output directory path
    output_dir = os.path.join(script_dir, "output")
    os.makedirs(output_dir, exist_ok=True)

    # Output Excel file path
    excel_file = os.path.join(output_dir, "transaction_summary.xlsx")

    include_item_id = True  # Set this to False to remove item ID from the result
    item_limit = 5  # Set the desired limit for most sold/bought items (None for no limit)

    most_sold_items = find_most_sold_items_per_day(log_file, include_id=include_item_id)
    most_bought_items = find_most_bought_items_per_day(log_file, include_id=include_item_id)
    most_sold_items_limited = {date: items[:item_limit] if item_limit else items for date, items in
                               most_sold_items.items()}
    most_bought_items_limited = {date: items[:item_limit] if item_limit else items for date, items in
                                most_bought_items.items()}

    buy_players_per_day = find_most_buy_players_per_day(log_file, player_limit=5)
    sell_players_per_day = find_most_sell_players_per_day(log_file, player_limit=5)

    sheet_data = {
        "Most Buy Players per Day": buy_players_per_day,
        "Most Sell Players per Day": sell_players_per_day
    }

    create_excel_file(excel_file, sheet_data, most_sold_items_limited, most_bought_items_limited, item_limit=item_limit)


if __name__ == "__main__":
    main()
